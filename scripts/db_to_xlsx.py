#!/usr/bin/env python3
"""
db_to_xlsx.py — Export PostgreSQL items table → Club_Gear_DB.xlsx

Usage:
    python scripts/db_to_xlsx.py
    python scripts/db_to_xlsx.py --out /tmp/gear_export.xlsx
    python scripts/db_to_xlsx.py --include-retired   # default: included

Reads DATABASE_URL from environment or .env file.
Output matches the Club_Gear_DB.xlsx format exactly so it can be
re-imported with xlsx_to_db.py.
"""

from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
import psycopg2.extras
import psycopg2
import os
import argparse
from datetime import datetime
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(__file__), "../backend/.env"))


HEADERS = [
    "tag",
    "name",
    "description",
    "locker",
    "status",
    "available",
    "manufactured_date",
    "condition_notes",
    "borrowed_by",   # human-readable alias for borrowed_by_email
]

COL_WIDTHS = [8, 22, 35, 12, 10, 10, 22, 30, 25]

STATUS_BG = {
    "active":   ("FFFFFF", "F0F5FF"),   # (even row, odd row)
    "retired":  ("FFE0E0", "FFCCCC"),
    "missing":  ("FFF8DC", "FFEAA0"),
}


def make_border():
    t = Side(style="thin", color="CCCCCC")
    return Border(left=t, right=t, top=t, bottom=t)


def write_sheet(ws, rows):
    bdr = make_border()

    # Header
    hfill = PatternFill("solid", fgColor="2E4057")
    hfont = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    halign = Alignment(horizontal="center", vertical="center")
    for ci, (h, w) in enumerate(zip(HEADERS, COL_WIDTHS), 1):
        c = ws.cell(1, ci, h)
        c.font = hfont
        c.fill = hfill
        c.alignment = halign
        c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22

    dfont = Font(name="Arial", size=9)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(vertical="center")
    CENTER_COLS = {"tag", "locker", "status", "available"}

    for ri, row in enumerate(rows, 2):
        s = row.get("status", "active")
        bg = STATUS_BG.get(s, ("FFFFFF", "F0F5FF"))[ri % 2]
        fill = PatternFill("solid", fgColor=bg)
        for ci, h in enumerate(HEADERS, 1):
            # borrowed_by in sheet = borrowed_by_email in db
            db_key = "borrowed_by_email" if h == "borrowed_by" else h
            val = row.get(db_key)
            if isinstance(val, bool):
                val = "TRUE" if val else "FALSE"
            elif val is None:
                val = ""
            c = ws.cell(ri, ci, val)
            c.font = dfont
            c.fill = fill
            c.border = bdr
            c.alignment = center if h in CENTER_COLS else left

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"


def write_legend(wb):
    ls = wb.create_sheet("legend")
    ldata = [
        ("Field", "Values / Notes"),
        ("tag", "item tag number e.g. 001"),
        ("name", "item type: harness, cam, quickdraw, helmet, etc."),
        ("description", "model/spec e.g. BD C4 size 1 red"),
        ("locker", "outdoor | top | bottom | pad | (blank = retired/missing)"),
        ("status", "active | retired | missing"),
        ("available", "TRUE = in locker; FALSE = on loan or unavailable"),
        ("manufactured_date", "year or date manufactured/purchased"),
        ("condition_notes", "free-text condition from stock check"),
        ("borrowed_by", "email of person currently borrowing; blank if in locker"),
        ("", ""),
        ("LOCKER VALUES", "DB column: items.locker"),
        ("outdoor", "T wall outer locker (most gear)"),
        ("top", "Top locker"),
        ("bottom", "Bottom locker"),
        ("pad", "Pad stash"),
        ("", ""),
        ("NOTE", "Edit status/available/borrowed_by then re-import with xlsx_to_db.py"),
        ("NOTE", "retired/missing rows: locker left blank intentionally"),
    ]
    bfont = Font(bold=True)
    for ri, (a, b) in enumerate(ldata, 1):
        ls.cell(ri, 1, a).font = bfont if ri == 1 else Font()
        ls.cell(ri, 2, b).font = bfont if ri == 1 else Font()
    ls.column_dimensions["A"].width = 28
    ls.column_dimensions["B"].width = 58


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--out", default="Club_Gear_DB.xlsx")
    args = parser.parse_args()

    db_url = os.getenv(
        "DATABASE_URL", "postgresql://gearuser:gearpass@localhost/gear")

    print("Connecting to database …")
    conn = psycopg2.connect(db_url)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    cur.execute("""
        SELECT
            tag, name, description, locker, status, available,
            manufactured_date, condition_notes, borrowed_by_email
        FROM items
        ORDER BY
            CASE status WHEN 'active' THEN 0 WHEN 'missing' THEN 1 ELSE 2 END,
            tag NULLS LAST
    """)
    rows = [dict(r) for r in cur.fetchall()]
    cur.close()
    conn.close()
    print(f"  {len(rows)} items fetched")

    wb = Workbook()
    ws = wb.active
    ws.title = "items"
    write_sheet(ws, rows)
    write_legend(wb)

    wb.save(args.out)
    print(f"  Saved → {args.out}")
    print("Done.")


if __name__ == "__main__":
    main()
